import openpyxl as xl
from openpyxl.chart import BarChart, Reference


#  loads the excel file
wb = xl.load_workbook('transactions.xlsx') 

#  get the active worksheet
sheet = wb['Sheet1']

# #  to access a cell 
# cell = sheet['a1']
# print(cell.value)

# cell = sheet.cell(1, 1)
# print(cell.value)

# # number of rows
# print(sheet.max_row)


def process_workbook(filename):
    wb = xl.load_workbook(filename) 
    sheet = wb['Sheet1']
    
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        new_price = cell.value * 0.9
        new_price_cell = sheet.cell(row, 4)
        new_price_cell.value = new_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


process_workbook('transactions.xlsx')